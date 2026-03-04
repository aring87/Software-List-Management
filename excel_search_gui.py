# edrl_excel_search_gui_v6.py
#
# FULL FILE (Search + Results + Request Queue + Request Form + Attachments)
#
# v6 updates:
# - Adds Queue Search (global across all queue columns)
# - Adds "Emergency Only" toggle on Queue tab
#
# Dependencies:
#   pip install pandas openpyxl
#
# Run:
#   python edrl_excel_search_gui_v6.py

import os
import sys
import re
import calendar
import json
import time
from PIL import Image, ImageTk
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd
from openpyxl import load_workbook


def resource_path(relative_path: str) -> str:
    base_path = getattr(sys, "_MEIPASS", os.path.abspath("."))
    return os.path.join(base_path, relative_path)


DEFAULT_WORKBOOK = resource_path("EDRL list v6.xlsx")

# Persist Request Form window geometry between runs
REQ_FORM_GEOM_FILE = resource_path("request_form_geometry.json")

BASE_FONT = ("Segoe UI", 12)
HEADER_FONT = ("Segoe UI", 18, "bold")
SUBHEADER_FONT = ("Segoe UI", 10)
BUTTON_FONT_BIG = ("Segoe UI", 14, "bold")
TREE_FONT = ("Segoe UI", 11)
TREE_HEADING_FONT = ("Segoe UI", 11, "bold")

BLUE_PRIMARY = "#1E5AA8"
LIGHT_BG = "#F3F8FF"
LIGHT_BG_2 = "#E6F0FF"
ROW_ODD = "#EAF2FF"
ROW_EVEN = "#FFFFFF"

NOISE_WORDS = {
    "installer", "setup", "client", "enterprise", "x64", "64bit", "64-bit", "x86", "msi", "win32",
    "machine", "wide", "machinewide", "update", "updater", "for", "windows", "mac", "osx", "macos",
    "app", "application", "software", "tool", "tools"
}


def norm_text(s: str) -> str:
    s = "" if s is None else str(s)
    s = s.lower().strip()
    s = re.sub(r"[^\w\s\.]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    tokens = [t for t in s.split(" ") if t and t not in NOISE_WORDS]
    return " ".join(tokens)


def extract_version_norm(v: str):
    v = "" if v is None else str(v).lower()
    cleaned = re.sub(r"[^0-9\.]", "", v)
    parts = [p for p in cleaned.split(".") if p != ""]
    nums = []
    for p in parts[:6]:
        try:
            nums.append(int(p))
        except Exception:
            nums.append(0)
    while len(nums) < 4:
        nums.append(0)
    return tuple(nums[:4])


def fmt_req_number(n: int) -> str:
    return f"REQ-{n:04d}"


def normalize_state(value: str) -> str:
    v = "" if value is None else str(value).strip().lower()
    if v in {"approved", "approve", "yes", "y", "true", "1"}:
        return "approved"
    if v in {"not approved", "notapproved", "no", "n", "false", "0", "disapproved", "denied", "reject", "rejected"}:
        return "not approved"
    if "not" in v and "approved" in v:
        return "not approved"
    if "approved" in v:
        return "approved"
    return ""


def open_file_with_default_app(path: str):
    if sys.platform.startswith("win"):
        os.startfile(path)  # type: ignore[attr-defined]
    elif sys.platform.startswith("darwin"):
        import subprocess
        subprocess.run(["open", path], check=False)
    else:
        import subprocess
        subprocess.run(["xdg-open", path], check=False)


def join_attachments(paths):
    clean = []
    for p in paths:
        if not p:
            continue
        p = str(p).strip()
        if p and p not in clean:
            clean.append(p)
    return " | ".join(clean)


def split_attachments(s: str):
    s = "" if s is None else str(s)
    parts = [p.strip() for p in s.split("|")]
    return [p for p in parts if p]


class DatePicker(tk.Toplevel):
    def __init__(self, parent, target_var: tk.StringVar, title="Select Date"):
        super().__init__(parent)
        self.title(title)
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        self.target_var = target_var

        today = pd.Timestamp.now()
        year = today.year
        month = today.month

        current = (target_var.get() or "").strip()
        try:
            if current:
                dt = pd.to_datetime(current, errors="raise")
                year = int(dt.year)
                month = int(dt.month)
        except Exception:
            pass

        self.year = tk.IntVar(value=year)
        self.month = tk.IntVar(value=month)

        top = ttk.Frame(self, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="Year").grid(row=0, column=0, sticky="w")
        ttk.Spinbox(top, from_=2000, to=2100, textvariable=self.year, width=8, command=self.refresh).grid(
            row=0, column=1, padx=(6, 14)
        )
        ttk.Label(top, text="Month").grid(row=0, column=2, sticky="w")
        ttk.Spinbox(top, from_=1, to=12, textvariable=self.month, width=5, command=self.refresh).grid(
            row=0, column=3, padx=(6, 0)
        )

        self.grid_frame = ttk.Frame(self, padding=(10, 0, 10, 10))
        self.grid_frame.pack()
        self.refresh()

    def refresh(self):
        for w in self.grid_frame.winfo_children():
            w.destroy()

        y = self.year.get()
        m = self.month.get()

        headers = ["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]
        for i, h in enumerate(headers):
            ttk.Label(self.grid_frame, text=h, width=4, anchor="center").grid(row=0, column=i)

        cal = calendar.monthcalendar(y, m)
        for r, week in enumerate(cal, start=1):
            for c, day in enumerate(week):
                if day == 0:
                    ttk.Label(self.grid_frame, text=" ", width=4).grid(row=r, column=c)
                else:
                    ttk.Button(
                        self.grid_frame,
                        text=str(day),
                        width=4,
                        command=lambda d=day: self.select_date(y, m, d),
                    ).grid(row=r, column=c, padx=1, pady=1)

    def select_date(self, y, m, d):
        self.target_var.set(f"{y:04d}-{m:02d}-{d:02d}")
        self.destroy()




class ScrollableFrame(ttk.Frame):
    """A vertical scrollable container for forms so fields never get cut off."""
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)

        # Prevent the form from becoming comically wide on large/zoomed windows.
        # We'll cap the *content* width and center it inside the canvas.
        self.max_inner_width = 1180

        self.canvas = tk.Canvas(self, highlightthickness=0, bd=0)
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.inner = ttk.Frame(self.canvas)
        self.inner_id = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.vsb.grid(row=0, column=1, sticky="ns")

        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)

        # Mousewheel support (Windows/macOS/Linux)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)       # Windows/macOS
        self.canvas.bind_all("<Button-4>", self._on_mousewheel_linux)   # Linux up
        self.canvas.bind_all("<Button-5>", self._on_mousewheel_linux)   # Linux down
        
        # Always start scrolled to the top after layout/scrollregion settles
        self.after(30, lambda: self.canvas.yview_moveto(0.0))
        self.after(150, lambda: self.canvas.yview_moveto(0.0))  # second tap for reliability
        
    def _on_inner_configure(self, _):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        # Resize the inner frame to the usable canvas width (DPI-safe)
        sbw = self.vsb.winfo_width()
        if sbw <= 1:
            sbw = 24  # fallback on first layout pass

        try:
            ht = int(self.canvas.cget("highlightthickness") or 0)
        except Exception:
            ht = 0
        try:
            bd = int(self.canvas.cget("bd") or 0)
        except Exception:
            bd = 0

        usable = max(1, event.width - sbw - (ht * 2) - (bd * 2) - 12)  # small safety gutter
        w = min(usable, self.max_inner_width)

        self.canvas.itemconfigure(self.inner_id, width=w)
        self.canvas.coords(self.inner_id, 0, 0)

    def _on_mousewheel(self, event):
        if self.winfo_containing(event.x_root, event.y_root) is None:
            return
        # Windows delta is 120 increments
        delta = -1 * int(event.delta / 120) if event.delta else 0
        if delta:
            self.canvas.yview_scroll(delta, "units")

    def _on_mousewheel_linux(self, event):
        if self.winfo_containing(event.x_root, event.y_root) is None:
            return
        if event.num == 4:
            self.canvas.yview_scroll(-3, "units")
        elif event.num == 5:
            self.canvas.yview_scroll(3, "units")


class EDRLSearchGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Software Search")
        self.root.geometry("1650x960")
        self.root.minsize(1250, 760)
        self.root.configure(bg=LIGHT_BG)

        try:
            self.root.option_add("*Font", f"{{{BASE_FONT[0]}}} {BASE_FONT[1]}")
        except Exception:
            pass

        self.root.rowconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)

        self.workbook_path = tk.StringVar(value=DEFAULT_WORKBOOK)

        self.sheets = {}
        self.idx_sheets = {}

        self.current_sheet = tk.StringVar(value="All")
        self.vendor_filter = tk.StringVar(value="(Any)")
        self.platform_filter = tk.StringVar(value="(Any)")
        self.status_filter = tk.StringVar(value="(Any)")
        self.state_filter = tk.StringVar(value="(Any)")
        self.search_field = tk.StringVar(value="All searchable fields")
        self.query_var = tk.StringVar(value="")
        self.name_query_var = tk.StringVar(value="")
        self.dedup_var = tk.BooleanVar(value=True)

        self.queue_query_var = tk.StringVar(value="")
        self.queue_emergency_only_var = tk.BooleanVar(value=False)

        self.platform_opts = ["Windows", "MacOS", "Cloud", "Mobile"]
        self.auth_user_opts = [
            "***OGC USE ONLY***",
            "***Restricted***Approved Resonable Accommodation Only",
            "***IT Cyber Use Only***",
            "***IT USE ONLY***",
            "DoDEA Staff",
            "DoDEA Students",
            "Medical",
            "All",
        ]
        self.type_opts = ["Cloud", "Software", "Extensions", "iOS", "Android"]
        self.yesno_opts = ["Yes", "No"]

        self.queue_columns_display = [
            "EDRL Number",
            "Name",
            "Version",
            "Type",
            "Platform",
            "Description",
            "Instructional Need",
            "Vendor",
            "Authorization Date",
            "Authorization Expiration",
            "Date Added",
            "State",
            "Authorized User",
            "Emergency",
            "URL",
            "Attachments",
            "Software Assessments",
        ]
        self._queue_internal_cols = ["_priority", "_created_ts"]

        self.queue_df = pd.DataFrame(columns=self.queue_columns_display + self._queue_internal_cols)

        self._apply_style()
        # Load DoDEA logos (two sizes)
        #self._dodea_logo_main = self._load_dodea_logo(max_height=500, cache_attr="_dodea_logo_main")
        #self._dodea_logo_req  = self._load_dodea_logo(max_height=175, cache_attr="_dodea_logo_req")
        self._build_ui()

        if os.path.exists(DEFAULT_WORKBOOK):
            self.load_workbook(DEFAULT_WORKBOOK)
        else:
            self.status_var.set("Default workbook not found. Click Browse… to select your workbook.")

    def _now_iso(self) -> str:
        return pd.Timestamp.now().isoformat(timespec="seconds")

    def _next_request_number(self) -> str:
        if self.queue_df.empty or "EDRL Number" not in self.queue_df.columns:
            return fmt_req_number(1)
        max_n = 0
        for v in self.queue_df["EDRL Number"].astype(str).tolist():
            m = re.match(r"^REQ-(\d+)$", v.strip(), flags=re.IGNORECASE)
            if m:
                try:
                    max_n = max(max_n, int(m.group(1)))
                except Exception:
                    pass
        return fmt_req_number(max_n + 1)

    # --- Request Form geometry persistence ---
    
    # ---------------- Logo helper ----------------
    def _load_dodea_logo(self, max_height: int, cache_attr: str = "_dodea_logo_img"):
        """Load DoDEA_Logo.png and resize to an exact max_height (pixel-accurate) using PIL. Cache on self."""
        try:
            path = resource_path("DoDEA_Logo.png")
            if not os.path.exists(path):
                setattr(self, cache_attr, None)
                return None

            img = Image.open(path).convert("RGBA")
            w, h = img.size
            if h <= 0 or not max_height or max_height <= 0:
                tk_img = ImageTk.PhotoImage(img)
                setattr(self, cache_attr, tk_img)
                return tk_img

            # Keep height fixed, force width wider (stretches horizontally)
            new_h = max_height
            new_w = 450  # <-- change this number to make it wider/narrower

            img = img.resize((new_w, new_h), Image.LANCZOS)
            tk_img = ImageTk.PhotoImage(img)
            setattr(self, cache_attr, tk_img)
            return tk_img
        except Exception:
            setattr(self, cache_attr, None)
            return None

    def _load_req_form_geometry(self) -> str | None:
        try:
            if os.path.exists(REQ_FORM_GEOM_FILE):
                with open(REQ_FORM_GEOM_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f) if f else {}
                geom = data.get("geometry")
                return geom if isinstance(geom, str) and "x" in geom else None
        except Exception:
            pass
        return None

    def _save_req_form_geometry(self, win: tk.Toplevel) -> None:
        try:
            with open(REQ_FORM_GEOM_FILE, "w", encoding="utf-8") as f:
                json.dump({"geometry": win.geometry()}, f)
        except Exception:
            pass

    def _apply_style(self):
        style = ttk.Style()
        try:
            if "clam" in style.theme_names():
                style.theme_use("clam")
        except Exception:
            pass

        # Global backgrounds
        style.configure("TFrame", background=LIGHT_BG)
        style.configure("TLabelframe", background=LIGHT_BG)
        style.configure("TLabelframe.Label", background=LIGHT_BG, font=("Segoe UI", 12, "bold"), foreground="#123B6F")
        style.configure("TLabel", background=LIGHT_BG, foreground="#0F2D57")
        style.configure("Status.TLabel", font=SUBHEADER_FONT, foreground="#234B7A", background=LIGHT_BG)

        # Make all entry / combobox fields WHITE (instead of the default grey)
        style.configure("TEntry", fieldbackground="white", background="white", foreground="#0F2D57")
        style.configure("TCombobox", fieldbackground="white", background="white", foreground="#0F2D57")
        style.map("TCombobox", fieldbackground=[("readonly", "white")], background=[("readonly", "white")])

        # Buttons + tree
        style.configure("Treeview", rowheight=30, font=TREE_FONT)
        style.configure("Treeview.Heading", font=TREE_HEADING_FONT, background="#2A6BC0", foreground="white")
        style.map("Treeview.Heading", background=[("active", "#174B8D")])

        style.configure(
            "Primary.TButton",
            font=("Segoe UI", 12, "bold"),
            padding=(10, 8),
            foreground="white",
            background=BLUE_PRIMARY,
        )
        style.map("Primary.TButton", background=[("active", "#174B8D"), ("pressed", "#123B6F")])

        # Dedup checkbox (theme)
        style.configure("Dedup.TCheckbutton", background=LIGHT_BG, foreground="#0F2D57")
        style.configure("Queue.TCheckbutton", background=LIGHT_BG, foreground=BLUE_PRIMARY, font=("Segoe UI", 11, "bold"))
        style.map("Queue.TCheckbutton", background=[("active", LIGHT_BG)], foreground=[("active", BLUE_PRIMARY)])
        style.map("Dedup.TCheckbutton", background=[("active", LIGHT_BG)])

        style.configure(
            "BigPrimary.TButton",
            font=BUTTON_FONT_BIG,
            padding=(18, 12),
            foreground="white",
            background=BLUE_PRIMARY,
        )
        style.map("BigPrimary.TButton", background=[("active", "#174B8D"), ("pressed", "#123B6F")])

        # Slightly larger "Request Software" button
        style.configure("HugePrimary.TButton", font=("Segoe UI", 15, "bold"), padding=(22, 14), foreground="white", background=BLUE_PRIMARY)
        style.map("HugePrimary.TButton", background=[("active", "#174B8D"), ("pressed", "#123B6F")])

        style.configure("Secondary.TButton", font=("Segoe UI", 12, "bold"), padding=(10, 8), foreground="white", background="#2A6BC0")
        style.map("Secondary.TButton", background=[("active", "#174B8D"), ("pressed", "#123B6F")])

        # Notebook styling (reduce default grey)
        style.configure("TNotebook", background=LIGHT_BG, borderwidth=0)
        style.configure("TNotebook.Tab", background=LIGHT_BG_2, foreground="#0F2D57", padding=(12, 8))
        style.map("TNotebook.Tab",
                  background=[("selected", BLUE_PRIMARY), ("active", "#2A6BC0")],
                  foreground=[("selected", "white"), ("active", "white")])

        # Separators
        style.configure("TSeparator", background=LIGHT_BG_2)

        # Scrollbars (more blue-friendly)
        style.configure("Vertical.TScrollbar", troughcolor=LIGHT_BG_2, background=BLUE_PRIMARY, bordercolor=LIGHT_BG_2, arrowcolor="#0F2D57")
        style.configure("Horizontal.TScrollbar", troughcolor=LIGHT_BG_2, background=BLUE_PRIMARY, bordercolor=LIGHT_BG_2, arrowcolor="#0F2D57")


    def _make_tree(self, parent):
        frame = ttk.Frame(parent)
        frame.pack(fill="both", expand=True)
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        tree = ttk.Treeview(frame, show="headings")
        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        tree.tag_configure("odd", background=ROW_ODD)
        tree.tag_configure("even", background=ROW_EVEN)
        return tree

    def _configure_tree_columns(self, tree: ttk.Treeview, columns):
        tree["columns"] = list(columns)
        for c in columns:
            tree.heading(c, text=c)
            tree.column(c, width=160, anchor="w", stretch=True)

    def _autosize_columns_to_content(self, tree: ttk.Treeview, max_rows_scan: int = 250):
        cols = list(tree["columns"])
        if not cols:
            return
        try:
            import tkinter.font as tkfont
            f = tkfont.Font(family=TREE_FONT[0], size=TREE_FONT[1])
        except Exception:
            f = None

        def measure(text: str) -> int:
            text = "" if text is None else str(text)
            return f.measure(text) if f else int(len(text) * 7)

        widths = {c: measure(c) + 28 for c in cols}
        kids = tree.get_children()[:max_rows_scan]
        for iid in kids:
            vals = tree.item(iid, "values")
            for c, v in zip(cols, vals):
                widths[c] = max(widths[c], measure(v) + 28)

        for c in cols:
            w = widths[c]
            cl = c.strip().lower()
            if cl in ("description", "instructional need"):
                w = min(max(w, 320), 950)
            elif cl == "url":
                w = min(max(w, 260), 650)
            elif cl in ("attachments", "software assessments"):
                w = min(max(w, 320), 950)
            else:
                w = min(max(w, 140), 480)
            tree.column(c, width=w, stretch=True)

    def _install_copy_shortcuts(self, tree: ttk.Treeview):
        def select_all(_=None):
            kids = tree.get_children()
            if kids:
                tree.selection_set(kids)

        def copy_rows(_=None):
            sel = tree.selection()
            if not sel:
                return
            cols = list(tree["columns"])
            lines = ["\t".join(cols)]
            for iid in sel:
                vals = tree.item(iid, "values")
                lines.append("\t".join("" if v is None else str(v) for v in vals))
            self.root.clipboard_clear()
            self.root.clipboard_append("\n".join(lines))

        tree.bind("<Control-a>", select_all)
        tree.bind("<Control-A>", select_all)
        tree.bind("<Control-c>", copy_rows)
        tree.bind("<Control-C>", copy_rows)

    def _find_col(self, df: pd.DataFrame, names_lower):
        cols = {str(c).strip().lower(): c for c in df.columns}
        for n in names_lower:
            if n in cols:
                return cols[n]
        return None

    def _find_sheet_case_insensitive(self, sheet_names, wanted: str):
        wanted_l = wanted.strip().lower()
        for s in sheet_names:
            if str(s).strip().lower() == wanted_l:
                return s
        return None

    def _build_ui(self):
        main = ttk.Frame(self.root, style='TFrame')
        main.grid(row=0, column=0, sticky="nsew")
        main.rowconfigure(2, weight=1)
        main.columnconfigure(0, weight=1)

        header = ttk.Frame(main, padding=(16, 14, 16, 8))
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(0, weight=1)

        ttk.Label(header, text="Software Search", font=HEADER_FONT).grid(row=0, column=0, sticky="w")
        ttk.Label(header, text="Ctrl+C copy • Ctrl+A select all", font=SUBHEADER_FONT, foreground="#3A5E8C").grid(
            row=1, column=0, sticky="w", pady=(2, 0)
        )

        wb_row = ttk.Frame(header)
        wb_row.grid(row=0, column=1, rowspan=2, sticky="e")

        ttk.Label(wb_row, text="Workbook:").grid(row=0, column=0, sticky="e", padx=(0, 6))
        ttk.Entry(wb_row, textvariable=self.workbook_path, width=64).grid(row=0, column=1, padx=(0, 10), sticky="we")
        ttk.Button(wb_row, text="Browse…", command=self.browse_workbook, style="Secondary.TButton").grid(row=0, column=2, padx=(0, 8))
        ttk.Button(wb_row, text="Load", command=self.load_current_workbook, style="Primary.TButton").grid(
            row=0, column=3, padx=(0, 10)
        )

        # ---- Controls row: left outlined "Search + Filters" + right "REQUEST SOFTWARE" (no outline) ----
        # UI goal:
        # - The outlined box should END after Search/Clear (no giant empty bordered area on wide windows).
        # - The Request button should live OUTSIDE the outline and be easy to nudge horizontally.
        controls_row = ttk.Frame(main)
        controls_row.grid(row=1, column=0, sticky="ew", padx=16, pady=(6, 10))
        # Lock the left two columns to content; let only the far-right spacer grow.
        controls_row.columnconfigure(0, weight=0)
        controls_row.columnconfigure(1, weight=0)
        controls_row.columnconfigure(2, weight=1)  # spacer eats extra width

        controls = ttk.LabelFrame(controls_row, text="Search + Filters", padding=(14, 12))
        # IMPORTANT: sticky="w" (NOT "ew") ensures the border never stretches across the window.
        controls.grid(row=0, column=0, sticky="w")

        request_area = ttk.Frame(controls_row)
        # sticky="n" keeps the Request button vertically aligned with the Search/Clear stack.
        request_area.grid(row=0, column=1, sticky="n", padx=(18, 0), pady=(2, 0))

        # Right-side banner area (use the flexible spacer column) — place DoDEA logo here
        logo_host = ttk.Frame(controls_row)
        logo_host.grid(row=0, column=2, sticky="nsew")
        # Manual placement knobs (0..1)
        LOGO_RELX = 0.50
        LOGO_RELY = 0.45
        #if getattr(self, "_dodea_logo_main", None):
        #    logo_lbl = ttk.Label(logo_host, image=self._dodea_logo_main, background=LIGHT_BG)
        #   logo_lbl.image = self._dodea_logo_main
        #    logo_lbl.place(relx=LOGO_RELX, rely=LOGO_RELY, anchor="center")

        # ---- Uniform spacing: build filter rows as equal-padding "field groups" ----
        # NOTE: these are purely UI tuning knobs.
        GROUP_PAD_X = 14
        GROUP_PAD_Y = 2
        LABEL_TO_WIDGET_PAD_Y = 4

        def add_field_group(parent, col, label_text, widget_factory, *, group_padx=GROUP_PAD_X, group_pady=GROUP_PAD_Y):
            g = ttk.Frame(parent)
            g.grid(row=0, column=col, padx=(0, group_padx), pady=(0, group_pady), sticky="w")
            ttk.Label(g, text=label_text).grid(row=0, column=0, sticky="w")
            w = widget_factory(g)
            w.grid(row=1, column=0, sticky="w", pady=(LABEL_TO_WIDGET_PAD_Y, 0))
            return w

        # Row 0: Name | Vendor | Platform | Approval | Expired
        filters_row = ttk.Frame(controls)
        filters_row.grid(row=0, column=0, sticky="w")

        self.name_entry = add_field_group(filters_row, 0, "Name", lambda p: ttk.Entry(p, textvariable=self.name_query_var, width=30))
        self.name_entry.bind("<Return>", lambda e: self.run_search())

        self.vendor_combo = add_field_group(filters_row, 1, "Vendor", lambda p: ttk.Combobox(p, state="readonly", width=26, textvariable=self.vendor_filter))
        self.vendor_combo.bind("<<ComboboxSelected>>", lambda e: self.run_search())

        self.platform_combo = add_field_group(filters_row, 2, "Platform", lambda p: ttk.Combobox(p, state="readonly", width=18, textvariable=self.platform_filter))
        self.platform_combo.bind("<<ComboboxSelected>>", lambda e: self.run_search())

        self.state_combo = add_field_group(filters_row, 3, "Approval", lambda p: ttk.Combobox(p, state="readonly", width=16, textvariable=self.state_filter))
        self.state_combo.bind("<<ComboboxSelected>>", lambda e: self.run_search())

        self.status_combo = add_field_group(
            filters_row,
            4,
            "Expired",
            lambda p: ttk.Combobox(p, state="readonly", width=16, textvariable=self.status_filter),
            group_padx=0,  # last group (no extra right padding)
        )
        self.status_combo["values"] = ["(Any)", "Expired Only", "Not Expired Only"]
        self.status_combo.bind("<<ComboboxSelected>>", lambda e: self.run_search())

        # Row 1: Search field + Query
        search_row = ttk.Frame(controls)
        search_row.grid(row=1, column=0, sticky="w", pady=(10, 0))

        self.field_combo = add_field_group(search_row, 0, "Search field", lambda p: ttk.Combobox(p, state="readonly", width=22, textvariable=self.search_field))

        self.query_entry = add_field_group(search_row, 1, "Query", lambda p: ttk.Entry(p, textvariable=self.query_var, width=44), group_padx=0)
        self.query_entry.bind("<Return>", lambda e: self.run_search())

        ttk.Checkbutton(
            controls,
            text="Dedup by Product (show latest version)",
            variable=self.dedup_var,
            command=self.run_search,
            style="Dedup.TCheckbutton",
        ).grid(row=2, column=0, pady=(10, 0), sticky="w")
        # Layout behavior inside the outlined box:
        # Keep everything tight and left-aligned. Do NOT add a stretching spacer inside the LabelFrame,
        # otherwise wide windows look like a huge empty bordered area.
        controls.columnconfigure(0, weight=0)  # filters do NOT stretch horizontally
        controls.columnconfigure(1, weight=0)  # Search/Clear stack


        # Action buttons area (right side)
        # Search/Clear stack goes in the RED area; Request Software goes in the GREEN area (far right).
        btns = ttk.Frame(controls)
        # Pad from the last dropdown to the action buttons for a clean separation.
        btns.grid(row=0, column=1, rowspan=2, padx=(18, 0), sticky="n")

        ttk.Button(btns, text="Search", command=self.run_search, style="Primary.TButton", width=14).pack(pady=(0, 6))
        ttk.Button(btns, text="Clear", command=self.clear_all, width=14, style="Primary.TButton").pack(pady=(0, 0))

        # ---- Request button (outside the outline) ----
        # Manual nudge control: change REQ_BTN_NUDGE_X to fine-tune placement.
        REQ_BTN_NUDGE_X = 0
        REQ_BTN_NUDGE_Y = 36
        
        self.request_btn_inline = ttk.Button(
            request_area, text="REQUEST SOFTWARE", command=self.open_request_form, style="HugePrimary.TButton"
        )
        
        self.request_btn_inline.grid(row=0, column=0, padx=(REQ_BTN_NUDGE_X, 0), pady=(REQ_BTN_NUDGE_Y, 0), sticky="w")

        body = ttk.Frame(main, padding=(16, 0, 16, 12))
        body.grid(row=2, column=0, sticky="nsew")
        body.rowconfigure(0, weight=1)
        body.columnconfigure(0, weight=1)

        self.notebook = ttk.Notebook(body)
        self.notebook.grid(row=0, column=0, sticky="nsew")

        self.results_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.results_tab, text="Results")

        top_status = ttk.Frame(self.results_tab, padding=(2, 8))
        top_status.pack(fill="x")
        self.status_var = tk.StringVar(value="Load a workbook to begin.")
        ttk.Label(top_status, textvariable=self.status_var, style="Status.TLabel").pack(anchor="w")

        # Results actions (UI-only)
        results_actions = ttk.Frame(self.results_tab, padding=(2, 0, 2, 8))
        results_actions.pack(fill="x")
        results_actions.columnconfigure(0, weight=1)

        self.results_delete_btn = ttk.Button(
            results_actions,
            text="Delete Selected",
            command=self.delete_selected_result,
            style="Secondary.TButton",
            state="disabled",
        )
        self.results_delete_btn.grid(row=0, column=1, sticky="e")

        self.results_clear_btn = ttk.Button(
            results_actions,
            text="Clear Results",
            command=self.clear_results_table,
            style="Secondary.TButton",
            state="disabled",
        )
        self.results_clear_btn.grid(row=0, column=2, padx=(10, 0), sticky="e")

        self.results_tree = self._make_tree(self.results_tab)
        self._install_copy_shortcuts(self.results_tree)

        # Enable/disable Results actions based on selection/content (UI-only)
        self.results_tree.bind("<<TreeviewSelect>>", lambda e: self._update_results_buttons_state())
        self.results_tree.bind("<Delete>", lambda e: self.delete_selected_result())
        self._update_results_buttons_state()

        self.queue_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.queue_tab, text="Request Queue")

        queue_top = ttk.Frame(self.queue_tab, padding=(2, 8))
        queue_top.pack(fill="x")
        queue_top.columnconfigure(0, weight=1)
        self.queue_status_var = tk.StringVar(value="No requests yet.")
        ttk.Label(queue_top, textvariable=self.queue_status_var, style="Status.TLabel").grid(row=0, column=0, sticky="w")

        queue_controls = ttk.Frame(self.queue_tab, padding=(2, 0, 2, 8))
        queue_controls.pack(fill="x")
        queue_controls.columnconfigure(1, weight=1)

        ttk.Label(queue_controls, text="Queue Search:").grid(row=0, column=0, sticky="w", padx=(0, 8))
        q_entry = ttk.Entry(queue_controls, textvariable=self.queue_query_var)
        q_entry.grid(row=0, column=1, sticky="ew")
        q_entry.bind("<Return>", lambda e: self.refresh_queue_table())
        ttk.Button(queue_controls, text="Apply", command=self.refresh_queue_table, style="Primary.TButton").grid(
            row=0, column=2, padx=(10, 0)
        )
        ttk.Button(queue_controls, text="Clear", command=self.clear_queue_filter, style="Primary.TButton", width=10).grid(row=0, column=3, padx=(10, 0))
        ttk.Checkbutton(
            queue_controls, text="Emergency Only", variable=self.queue_emergency_only_var, command=self.refresh_queue_table, style="Queue.TCheckbutton"
        ).grid(row=0, column=4, padx=(14, 0), sticky="w")

        action_row = ttk.Frame(self.queue_tab, padding=(2, 0, 2, 8))
        action_row.pack(fill="x")
        action_row.columnconfigure(0, weight=1)

        self.edit_req_btn = ttk.Button(action_row, text="Edit", command=self.edit_selected_request, style="Secondary.TButton")
        self.edit_req_btn.grid(row=0, column=1, padx=(0, 10), sticky="e")

        self.delete_req_btn = ttk.Button(action_row, text="Delete", command=self.delete_selected_request, style="Secondary.TButton")
        self.delete_req_btn.grid(row=0, column=2, padx=(0, 10), sticky="e")

        self.add_software_btn = ttk.Button(action_row, text="Add Software", command=self.add_selected_request_to_all, style="Primary.TButton")
        self.add_software_btn.grid(row=0, column=3, padx=(0, 10), sticky="e")

        ttk.Button(action_row, text="Export Queue (EDRL Format)", command=self.export_queue_edrl, style="Secondary.TButton").grid(
            row=0, column=4, padx=(8, 0), sticky="e"
        )

        self.queue_tree = self._make_tree(self.queue_tab)
        self._install_copy_shortcuts(self.queue_tree)
        self._configure_tree_columns(self.queue_tree, self.queue_columns_display)

        self.queue_tree.bind("<<TreeviewSelect>>", lambda e: self._update_queue_buttons_state())
        self._update_queue_buttons_state()

    # workbook
    def browse_workbook(self):
        path = filedialog.askopenfilename(title="Select workbook", filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if path:
            self.workbook_path.set(path)

    def load_current_workbook(self):
        path = (self.workbook_path.get() or "").strip().strip('"')
        if not path:
            messagebox.showerror("Error", "Please select a workbook.")
            return
        self.load_workbook(path)

    def load_workbook(self, path: str):
        try:
            xls = pd.ExcelFile(path)
            sheet_names = xls.sheet_names
            self.sheets = {s: pd.read_excel(path, sheet_name=s, dtype=str).fillna("") for s in sheet_names}
            self.idx_sheets = {s: self.build_index(df) for s, df in self.sheets.items()}

            all_sheet = self._find_sheet_case_insensitive(sheet_names, "All") or sheet_names[0]
            preferred = []
            for candidate in ["All", "Cloud", "Chrome_AppExt", "Software", "Mobile", "Everything Else"]:
                real = self._find_sheet_case_insensitive(sheet_names, candidate)
                if real and real not in preferred:
                    preferred.append(real)
            ordered = preferred + [s for s in sheet_names if s not in preferred]

            if hasattr(self, "sheet_combo"):
                self.sheet_combo["values"] = ordered
            self.current_sheet.set(all_sheet)
            self.on_sheet_change()
            self.status_var.set(f"Loaded: {os.path.basename(path)} • Tabs: {len(sheet_names)}")
        except Exception as e:
            messagebox.showerror("Load failed", f"Could not load workbook.\n\n{e}")

    def build_index(self, df: pd.DataFrame) -> pd.DataFrame:
        idx = df.copy()
        idx["_blob"] = df.astype(str).agg(" ".join, axis=1).map(norm_text)

        exp_col = self._find_col(df, ["authorization expires", "authorization expiration", "authorization expiration date", "authorization expires date", "authorization expiration (date)"])
        idx["is_expired"] = False
        if exp_col:
            exp_dt = pd.to_datetime(df[exp_col], errors="coerce")
            now = pd.Timestamp.now().normalize()
            idx["is_expired"] = exp_dt.notna() & (exp_dt < now)

        name_col = self._find_col(df, ["name"])
        vendor_col = self._find_col(df, ["vendor"])
        platform_col = self._find_col(df, ["platform"])
        version_col = self._find_col(df, ["version"])

        name_s = df[name_col].astype(str) if name_col else pd.Series([""] * len(df))
        vendor_s = df[vendor_col].astype(str) if vendor_col else pd.Series([""] * len(df))
        platform_s = df[platform_col].astype(str) if platform_col else pd.Series([""] * len(df))
        version_s = df[version_col].astype(str) if version_col else pd.Series([""] * len(df))

        idx["name_norm"] = name_s.map(norm_text)
        idx["vendor_norm"] = vendor_s.map(norm_text)
        idx["platform_norm"] = platform_s.map(norm_text)
        idx["version_norm"] = version_s.map(extract_version_norm)
        idx["product_key"] = idx["vendor_norm"] + "|" + idx["name_norm"] + "|" + idx["platform_norm"]

        state_col = self._find_col(df, ["state"])
        idx["state_norm"] = ""
        if state_col:
            idx["state_norm"] = df[state_col].astype(str).map(normalize_state)
        return idx

    def dedup_latest(self, idx: pd.DataFrame, original: pd.DataFrame) -> pd.DataFrame:
        if "product_key" not in idx.columns or "version_norm" not in idx.columns:
            return original
        idx_sorted = idx.sort_values(["product_key", "version_norm"], ascending=[True, False])
        top = idx_sorted.groupby("product_key", as_index=False).head(1)
        return original.loc[top.index].copy()

    # search
    def on_sheet_change(self):
        sheet = self.current_sheet.get()
        if sheet not in self.sheets:
            return
        df = self.sheets[sheet]
        idx = self.idx_sheets[sheet]

        self.field_combo["values"] = ["All searchable fields"] + list(df.columns)
        self.search_field.set("All searchable fields")

        vendor_col = self._find_col(df, ["vendor"])
        platform_col = self._find_col(df, ["platform"])
        state_col = self._find_col(df, ["state"])

        self.vendor_combo["values"] = ["(Any)"] + (sorted({str(v).strip() for v in df[vendor_col].astype(str).tolist() if str(v).strip()}) if vendor_col else [])
        self.platform_combo["values"] = ["(Any)"] + (sorted({str(p).strip() for p in df[platform_col].astype(str).tolist() if str(p).strip()}) if platform_col else [])

        if state_col and "state_norm" in idx.columns:
            present = set([s for s in idx["state_norm"].astype(str).tolist() if s])
            vals = ["(Any)"]
            if "approved" in present:
                vals.append("Approved")
            if "not approved" in present:
                vals.append("Not Approved")
            if vals == ["(Any)"]:
                vals = ["(Any)", "Approved", "Not Approved"]
            self.state_combo["values"] = vals
        else:
            self.state_combo["values"] = ["(Any)"]

        self.vendor_filter.set("(Any)")
        self.platform_filter.set("(Any)")
        self.status_filter.set("(Any)")
        self.state_filter.set("(Any)")
        self.run_search()

    def run_search(self):
        sheet = self.current_sheet.get()
        if sheet not in self.sheets:
            return
        df = self.sheets[sheet]
        idx = self.idx_sheets[sheet]

        query = (self.query_var.get() or "").strip()
        name_query = (self.name_query_var.get() or "").strip()

        vendor_sel = self.vendor_filter.get()
        platform_sel = self.platform_filter.get()
        status_sel = self.status_filter.get()
        state_sel = self.state_filter.get()
        field_sel = self.search_field.get()

        vendor_col = self._find_col(df, ["vendor"])
        platform_col = self._find_col(df, ["platform"])
        name_col = self._find_col(df, ["name"])

        mask = pd.Series(True, index=df.index)

        if vendor_col and vendor_sel != "(Any)":
            mask &= df[vendor_col].astype(str).str.strip().eq(vendor_sel)
        if platform_col and platform_sel != "(Any)":
            mask &= df[platform_col].astype(str).str.strip().eq(platform_sel)

        if state_sel in {"Approved", "Not Approved"} and "state_norm" in idx.columns:
            want = "approved" if state_sel == "Approved" else "not approved"
            mask &= idx["state_norm"].astype(str).str.strip().eq(want)

        if status_sel == "Expired Only":
            mask &= idx["is_expired"].astype(bool)
        elif status_sel == "Not Expired Only":
            mask &= ~idx["is_expired"].astype(bool)

        if name_query:
            if name_col:
                mask &= df[name_col].astype(str).str.contains(name_query, case=False, na=False)
            else:
                mask &= idx["_blob"].astype(str).str.contains(norm_text(name_query), na=False)

        if query:
            if field_sel == "All searchable fields":
                mask &= idx["_blob"].astype(str).str.contains(norm_text(query), na=False)
            elif field_sel in df.columns:
                mask &= df[field_sel].astype(str).str.contains(query, case=False, na=False)

        results = df.loc[mask].copy()
        if self.dedup_var.get():
            results = self.dedup_latest(idx.loc[mask], results)
        self.populate_results_table(results)

    def clear_all(self):
        self.query_var.set("")
        self.name_query_var.set("")
        self.vendor_filter.set("(Any)")
        self.platform_filter.set("(Any)")
        self.status_filter.set("(Any)")
        self.state_filter.set("(Any)")
        self.search_field.set("All searchable fields")
        self.dedup_var.set(True)
        self.run_search()

    def populate_results_table(self, df: pd.DataFrame):
        self.results_tree.delete(*self.results_tree.get_children())
        self._configure_tree_columns(self.results_tree, df.columns)
        for i, row in enumerate(df.head(6000).itertuples(index=False, name=None)):
            self.results_tree.insert("", "end", values=row, tags=("even" if i % 2 == 0 else "odd"))
        self._autosize_columns_to_content(self.results_tree, max_rows_scan=250)

    # queue
    def clear_queue_filter(self):
        self.queue_query_var.set("")
        self.queue_emergency_only_var.set(False)
        self.refresh_queue_table()

    def _queue_filtered_sorted_df(self) -> pd.DataFrame:
        df = self.queue_df.copy()

        # Emergency goes to the top (Yes first), then oldest first.
        if "Emergency" in df.columns:
            em = df["Emergency"].astype(str).str.strip().str.lower()
            df["_priority"] = em.ne("yes").astype(int)  # yes -> 0 (top), else -> 1
        else:
            df["_priority"] = 1

        if "_created_ts" not in df.columns:
            df["_created_ts"] = list(range(len(df)))
        else:
            # fillna() does not accept a list; use an index-aligned Series as fallback
            tmp = pd.to_numeric(df["_created_ts"], errors="coerce")
            df["_created_ts"] = tmp.fillna(pd.Series(range(len(df)), index=df.index))

        if self.queue_emergency_only_var.get() and "Emergency" in df.columns:
            df = df[df["Emergency"].astype(str).str.strip().str.lower().eq("yes")].copy()

        q = (self.queue_query_var.get() or "").strip()
        if q:
            q_norm = q.lower()
            cols = [c for c in self.queue_columns_display if c in df.columns]
            blob = df[cols].astype(str).agg(" ".join, axis=1).str.lower()
            df = df[blob.str.contains(re.escape(q_norm), na=False)].copy()

        df = df.sort_values(by=["_priority", "_created_ts"], ascending=[True, True], kind="mergesort")
        return df

    def refresh_queue_table(self):
        df = self._queue_filtered_sorted_df()
        self.queue_tree.delete(*self.queue_tree.get_children())
        for i, row in enumerate(df[self.queue_columns_display].itertuples(index=False, name=None)):
            self.queue_tree.insert("", "end", values=row, tags=("even" if i % 2 == 0 else "odd"))
        total = len(self.queue_df)
        shown = len(df)
        suffix = f" (showing {shown} of {total})" if ((self.queue_query_var.get() or "").strip() or self.queue_emergency_only_var.get()) else ""
        self.queue_status_var.set(f"Requests in queue: {total}{suffix}")
        self._autosize_columns_to_content(self.queue_tree, max_rows_scan=250)

    # simple request form (kept minimal to focus on queue-search feature)
        # -------------------------- Request Form --------------------------

    # -------------------------- Request Form --------------------------
    def open_request_form(self, prefill: dict | None = None, edit_req_number: str | None = None):
        """Request Software form (v4 layout) styled to match the v6 theme."""
        
        prefill = prefill or {}
        is_edit = bool(edit_req_number)
        win = tk.Toplevel(self.root)
        win.title("Request Software")

        # Start large (taskbar-safe). We will auto-fit width again after building the form.
        win.minsize(940, 780)
        win.transient(self.root)
        win.grab_set()

        # Restore last size/position if available (but we will still enforce a minimum after building)
        saved_geom = self._load_req_form_geometry()
        if saved_geom:
            try:
                win.geometry(saved_geom)
            except Exception:
                pass

        def _close_req_form():
            # Save window geometry and close cleanly (important when using grab_set).
            try:
                self._save_req_form_geometry(win)
            except Exception:
                pass
            try:
                win.grab_release()
            except Exception:
                pass
            try:
                win.destroy()
            except Exception:
                pass

        win.protocol("WM_DELETE_WINDOW", _close_req_form)

        # ----- Theme (match main app) -----
        BG_APP = "#F6F9FF"
        BG_PANEL = "#FFFFFF"
        BORDER = "#CFE0FF"
        TEXT_DARK = "#0F172A"
        MUTED = "#475569"
        BLUE_PRIMARY = "#1E5AA8"
        BLUE_HOVER = "#174A8A"
        ORANGE_EMERGENCY = "#C2410C"

        win.configure(bg=BG_APP)

        # DoDEA logo in the top-right blank header area (manual knobs)
        REQ_LOGO_RELX = 0.87
        REQ_LOGO_RELY = 0.02
        #if getattr(self, "_dodea_logo_req", None):
        #    try:
        #        win._logo_ref = self._dodea_logo_req  # prevent GC
        #        logo_lbl = tk.Label(win, image=self._dodea_logo_req, bg=BG_APP, bd=0, highlightthickness=0)
        #        logo_lbl.image = self._dodea_logo_req
        #       logo_lbl.place(relx=REQ_LOGO_RELX, rely=REQ_LOGO_RELY, anchor="ne")
        #        win.after(80, logo_lbl.lift)
        #    except Exception:
        #        pass

        style = ttk.Style(win)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        # Containers / text
        style.configure("Req.App.TFrame", background=BG_APP)
        style.configure("Req.Card.TFrame", background=BG_PANEL, relief="solid", borderwidth=1)
        style.configure("Req.Title.TLabel", background=BG_APP, foreground=TEXT_DARK, font=("Segoe UI", 15, "bold"))
        style.configure("Req.Subtitle.TLabel", background=BG_APP, foreground=MUTED, font=("Segoe UI", 9))

        style.configure("Req.Field.TLabel", background=BG_APP, foreground=TEXT_DARK, font=("Segoe UI", 10))
        style.configure("Req.Sub.TLabel", background=BG_APP, foreground="#64748B", font=("Segoe UI", 9))

        # Inputs
        style.configure("Req.TEntry", padding=5)
        style.configure("Req.TCombobox", padding=3)

        # Buttons (blue, consistent sizing)
        style.configure("Req.Primary.TButton", font=("Segoe UI", 11, "bold"), padding=(18, 8),
                        foreground="white", background=BLUE_PRIMARY)
        style.map("Req.Primary.TButton",
                  background=[("active", BLUE_HOVER), ("pressed", "#123C72")],
                  foreground=[("disabled", "#E6EEF8")])
        style.configure("Req.Action.TButton", font=("Segoe UI", 10, "bold"), padding=(14, 7),
                        foreground="white", background=BLUE_PRIMARY)
        style.map("Req.Action.TButton",
                  background=[("active", BLUE_HOVER), ("pressed", "#123C72")])

        # Date "Pick" buttons (same look/height as other blue action buttons)
        style.configure("Req.Pick.TButton", font=("Segoe UI", 10, "bold"), padding=(14, 7),
                        foreground="white", background=BLUE_PRIMARY)
        style.map("Req.Pick.TButton",
                  background=[("active", BLUE_HOVER), ("pressed", "#123C72")])

        # Emergency checkbox label
        style.configure("Req.Emergency.TCheckbutton", background=BG_APP, foreground=ORANGE_EMERGENCY,
                        font=("Segoe UI", 10, "bold"))

        outer = ttk.Frame(win, padding=14, style="Req.App.TFrame")
        outer.pack(fill="both", expand=True)
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(2, weight=1)

        ttk.Label(outer, text="Request Software", style="Req.Title.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(
            outer,
            text="Mandatory: Date, Platform, Authorized User, Type, Product Name, Description, Justification",
            style="Req.Subtitle.TLabel",
        ).grid(row=1, column=0, sticky="w", pady=(2, 10))

        # Scrollable body (caps width so the form doesn't stretch across huge monitors)
        body = ScrollableFrame(outer)
        body.grid(row=2, column=0, sticky="nsew")
        body.configure(style="Req.App.TFrame")
        body.canvas.configure(background=BG_APP)
        form = body.inner
        form.configure(style="Req.App.TFrame")
        
        # Grid columns: label | field-area (most rows live in col=1 and optionally span)
        form.columnconfigure(0, weight=0)
        form.columnconfigure(1, weight=0)
        form.columnconfigure(2, weight=0)
        form.columnconfigure(3, weight=0)

        # Vars
        self._req_form_first_focus = None
        number_preview = tk.StringVar(value=str(prefill.get("EDRL Number") or self._next_request_number()))

        date_var = tk.StringVar(value=str(prefill.get("Date Added") or prefill.get("Date") or ""))
        platform_var = tk.StringVar(value=str(prefill.get("Platform") or "Windows"))
        auth_user_var = tk.StringVar(value=str(prefill.get("Authorized User") or "DoDEA Staff"))
        emergency_var = tk.BooleanVar(value=str(prefill.get("Emergency") or "").strip().lower() in ("yes","y","true","1"))
        type_var = tk.StringVar(value=str(prefill.get("Type") or "Software"))
        version_var = tk.StringVar(value=str(prefill.get("Version") or ""))
        vendor_var = tk.StringVar(value=str(prefill.get("Vendor") or ""))
        url_var = tk.StringVar(value=str(prefill.get("URL") or ""))
        customer_attach_var = tk.StringVar(value=str(prefill.get("Attachments") or ""))

        cyber_date_created = tk.StringVar(value="")
        cyber_approved_date = tk.StringVar(value="")
        cyber_disapproved_date = tk.StringVar(value="")
        cyber_updated_date = tk.StringVar(value="")
        cyber_contract_cor = tk.StringVar(value="")
        cyber_auth_exp = tk.StringVar(value="")
        approved_yesno = tk.StringVar(value="No")
        cyber_assess_pdf_var = tk.StringVar(value=str(prefill.get("Software Assessments") or ""))

        def on_platform_change(*_):
            p = (platform_var.get() or "").strip()
            cur = (type_var.get() or "").strip()
            if p == "Cloud" and cur != "Cloud":
                type_var.set("Cloud")
            elif p == "Mobile" and cur not in ("iOS", "Android"):
                type_var.set("iOS")
            elif p in ("Windows", "MacOS") and cur not in ("Software", "Extensions"):
                type_var.set("Software")

        try:
            platform_var.trace_add("write", on_platform_change)
        except Exception:
            pass
        on_platform_change()

        # ---- Layout (v4 grid) ----
        r = 0
        ttk.Label(form, text="Number", style="Req.Field.TLabel").grid(row=r, column=0, sticky="w", pady=4)

        num_row = ttk.Frame(form, style="Req.App.TFrame")
        num_row.grid(row=r, column=1, columnspan=3, sticky="w", pady=4)
        ttk.Entry(num_row, textvariable=number_preview, width=18, state="readonly", style="Req.TEntry").pack(side="left")
        ttk.Label(num_row, text="(auto-generated on submit)", style="Req.Sub.TLabel").pack(side="left", padx=(10, 0))
        r += 1

        r = self._date_row_form(form, r, "Date *", date_var)

        ttk.Label(form, text="Platform *", style="Req.Field.TLabel").grid(row=r, column=0, sticky="w", pady=4)
        ttk.Combobox(
            form, textvariable=platform_var, values=self.platform_opts,
            state="readonly", width=18, style="Req.TCombobox"
        ).grid(row=r, column=1, sticky="w", pady=4)
        r += 1

        ttk.Label(form, text="Authorized User *", style="Req.Field.TLabel").grid(row=r, column=0, sticky="w", pady=4)
        au_row = ttk.Frame(form, style="Req.App.TFrame")
        au_row.grid(row=r, column=1, columnspan=3, sticky="w", pady=4)
        ttk.Combobox(
            au_row, textvariable=auth_user_var, values=self.auth_user_opts,
            state="readonly", width=72, style="Req.TCombobox"
        ).pack(side="left")
        ttk.Checkbutton(
            au_row, text="Emergency", variable=emergency_var, style="Req.Emergency.TCheckbutton"
        ).pack(side="left", padx=(10, 0))
        r += 1

        ttk.Label(form, text="Product Name *", style="Req.Field.TLabel").grid(row=r, column=0, sticky="nw", pady=4)
        product_txt = tk.Text(form, height=3, wrap="word", bd=1, relief="solid")
        product_txt.configure(width=96)
        product_txt.grid(row=r, column=1, columnspan=3, sticky="w", pady=4)
        try:
            product_txt.insert("1.0", str(prefill.get("Name") or ""))
        except Exception:
            pass
        r += 1

        ttk.Label(form, text="Type *", style="Req.Field.TLabel").grid(row=r, column=0, sticky="w", pady=4)
        ttk.Combobox(
            form, textvariable=type_var, values=self.type_opts,
            state="readonly", width=18, style="Req.TCombobox"
        ).grid(row=r, column=1, sticky="w", pady=4)
        r += 1

        ttk.Label(form, text="Description *", style="Req.Field.TLabel").grid(row=r, column=0, sticky="nw", pady=4)
        desc_txt = tk.Text(form, height=7, wrap="word", bd=1, relief="solid")
        desc_txt.configure(width=96)
        desc_txt.grid(row=r, column=1, columnspan=3, sticky="w", pady=4)
        try:
            desc_txt.insert("1.0", str(prefill.get("Description") or ""))
        except Exception:
            pass
        r += 1

        ttk.Label(form, text="Justification *", style="Req.Field.TLabel").grid(row=r, column=0, sticky="nw", pady=4)
        just_txt = tk.Text(form, height=5, wrap="word", bd=1, relief="solid")
        just_txt.configure(width=96)
        just_txt.grid(row=r, column=1, columnspan=3, sticky="w", pady=4)
        try:
            just_txt.insert("1.0", str(prefill.get("Justification") or prefill.get("Instructional Need") or ""))
        except Exception:
            pass
        r += 1

        ttk.Label(form, text="Version", style="Req.Field.TLabel").grid(row=r, column=0, sticky="w", pady=4)
        ttk.Entry(form, textvariable=version_var, width=30, style="Req.TEntry").grid(row=r, column=1, sticky="w", pady=4)
        r += 1

        ttk.Label(form, text="Vendor", style="Req.Field.TLabel").grid(row=r, column=0, sticky="w", pady=4)
        ttk.Entry(form, textvariable=vendor_var, width=46, style="Req.TEntry").grid(row=r, column=1, sticky="w", pady=4)
        r += 1

        ttk.Label(form, text="URL", style="Req.Field.TLabel").grid(row=r, column=0, sticky="w", pady=4)
        ttk.Entry(form, textvariable=url_var, width=100, style="Req.TEntry").grid(row=r, column=1, columnspan=3, sticky="w", pady=4)
        r += 1

        # CUSTOMER attachments (any files)
        ttk.Label(form, text="Attachments (docs/exe/iso/etc)", style="Req.Field.TLabel").grid(row=r, column=0, sticky="w", pady=4)

        att_row = ttk.Frame(form, style="Req.App.TFrame")
        att_row.grid(row=r, column=1, columnspan=3, sticky="we", pady=4)

        entry = ttk.Entry(att_row, textvariable=customer_attach_var, style="Req.TEntry")
        entry.pack(side="left", fill="x", expand=True)

        ttk.Button(att_row, text="Open…", style="Req.Action.TButton",
                   command=lambda: self._open_first_attachment(customer_attach_var.get())).pack(side="right", padx=(8, 0))
        ttk.Button(att_row, text="Add…", style="Req.Action.TButton",
                   command=lambda: self._pick_attachments_any(customer_attach_var)).pack(side="right", padx=(8, 0))

        r += 1

        ttk.Separator(form, orient="horizontal").grid(row=r, column=0, columnspan=4, sticky="we", pady=(18, 10))
        r += 1
        ttk.Label(form, text="*** CYBER STAFF ONLY ***", font=("Segoe UI", 10, "bold"), style="Req.Field.TLabel").grid(
            row=r, column=0, columnspan=4, sticky="w", pady=(0, 10)
        )
        r += 1

        r = self._date_row_form(form, r, "Date Created", cyber_date_created)
        r = self._date_row_form(form, r, "Approved Date", cyber_approved_date)
        r = self._date_row_form(form, r, "Disapproved Date", cyber_disapproved_date)
        r = self._date_row_form(form, r, "Updated", cyber_updated_date)

        ttk.Label(form, text="Contract COR", style="Req.Field.TLabel").grid(row=r, column=0, sticky="w", pady=4)
        ttk.Entry(form, textvariable=cyber_contract_cor, width=40, style="Req.TEntry").grid(row=r, column=1, sticky="w", pady=4)
        r += 1

        r = self._date_row_form(form, r, "Authorization Expiration", cyber_auth_exp)

        ttk.Label(form, text="Approved (Yes/No)", style="Req.Field.TLabel").grid(row=r, column=0, sticky="w", pady=4)
        ttk.Combobox(
            form, textvariable=approved_yesno, values=self.yesno_opts,
            state="readonly", width=10, style="Req.TCombobox"
        ).grid(row=r, column=1, sticky="w", pady=4)
        r += 1

        # CYBER attachments (PDF only)
        ttk.Label(form, text="Software Assessments (PDF only)", style="Req.Field.TLabel").grid(row=r, column=0, sticky="w", pady=4)

        pdf_row = ttk.Frame(form, style="Req.App.TFrame")
        pdf_row.grid(row=r, column=1, columnspan=3, sticky="we", pady=4)

        entry = ttk.Entry(pdf_row, textvariable=cyber_assess_pdf_var, style="Req.TEntry")
        entry.pack(side="left", fill="x", expand=True)

        ttk.Button(pdf_row, text="Open…", style="Req.Action.TButton",
                   command=lambda: self._open_first_attachment(cyber_assess_pdf_var.get())).pack(side="right", padx=(8, 0))
        ttk.Button(pdf_row, text="Add…", style="Req.Action.TButton",
                   command=lambda: self._pick_attachments_pdf_only(cyber_assess_pdf_var)).pack(side="right", padx=(8, 0))

        r += 1

        # --- Submit/Cancel aligned under Add/Open (right side of the attachment rows) ---
        ttk.Separator(form, orient="horizontal").grid(row=r, column=0, columnspan=4, sticky="we", pady=(14, 10))
        r += 1

        def on_submit():
            try:
                # mandatory fields
                if not (date_var.get() or "").strip():
                    messagebox.showerror("Missing info", "Date is required.")
                    return
                if not (platform_var.get() or "").strip():
                    messagebox.showerror("Missing info", "Platform is required.")
                    return
                if not (auth_user_var.get() or "").strip():
                    messagebox.showerror("Missing info", "Authorized User is required.")
                    return
                if not (type_var.get() or "").strip():
                    messagebox.showerror("Missing info", "Type is required.")
                    return
                
                product_name = product_txt.get("1.0", "end").strip()
                description = desc_txt.get("1.0", "end").strip()
                justification = just_txt.get("1.0", "end").strip()
                
                if not product_name:
                    messagebox.showerror("Missing info", "Product Name is required.")
                    return
                if not description:
                    messagebox.showerror("Missing info", "Description is required.")
                    return
                if not justification:
                    messagebox.showerror("Missing info", "Justification is required.")
                    return
                
                number_val = str(edit_req_number) if edit_req_number else self._next_request_number()
                
                record = {
                    "EDRL Number": number_val,
                    "Name": product_name,
                    "Version": (version_var.get() or "").strip(),
                    "Type": (type_var.get() or "").strip(),
                    "Platform": (platform_var.get() or "").strip(),
                    "Description": description,
                    "Justification": justification,
                    "Vendor": (vendor_var.get() or "").strip(),
                    "Authorization Date": (cyber_approved_date.get() or "").strip(),
                    "Authorization Expiration": (cyber_auth_exp.get() or "").strip(),
                    "Date Added": (cyber_date_created.get() or "").strip(),
                    "State": (approved_yesno.get() or "").strip(),
                    "Authorized User": (auth_user_var.get() or "").strip(),
                    "URL": (url_var.get() or "").strip(),
                    "Attachments": (customer_attach_var.get() or "").strip(),
                    "Software Assessments": (cyber_assess_pdf_var.get() or "").strip(),
                    "Emergency": ("Yes" if emergency_var.get() else "No"),
                }
                
                # Internal sort fields for queue ordering
                record["_created_ts"] = time.time()
                record["_priority"] = 0 if emergency_var.get() else 1
                
                
                # If your queue_df has extra columns, fill them safely
                if hasattr(self, "queue_columns"):
                    for c in self.queue_columns:
                        record.setdefault(c, "")
                
                # Update existing request if editing; otherwise append new row
                if edit_req_number and "EDRL Number" in self.queue_df.columns:
                    # Preserve internal columns if present
                    mask = self.queue_df["EDRL Number"].astype(str).str.strip() == str(edit_req_number).strip()
                    if mask.any():
                        for col in self.queue_df.columns:
                            if col in record:
                                self.queue_df.loc[mask, col] = record[col]
                        # keep internal cols as-is
                    else:
                        self.queue_df = pd.concat([self.queue_df, pd.DataFrame([record])], ignore_index=True)
                else:
                    self.queue_df = pd.concat([self.queue_df, pd.DataFrame([record])], ignore_index=True)
                if hasattr(self, "refresh_queue_table"):
                    self.refresh_queue_table()
                if hasattr(self, "notebook") and hasattr(self, "queue_tab"):
                    try:
                        self.notebook.select(self.queue_tab)
                    except Exception:
                        pass
                _close_req_form()
                
            except Exception as e:
                import traceback
                traceback.print_exc()
                messagebox.showerror('Submit failed', f'An unexpected error occurred while submitting.\n\n{e}')
        btn_row = ttk.Frame(form, style="Req.App.TFrame")
        btn_row.grid(row=r, column=1, columnspan=3, sticky="we", pady=(0, 6))

        # Stretch spacer: always pushes buttons to the right without clipping
        spacer = ttk.Frame(btn_row, style="Req.App.TFrame")
        spacer.pack(side="left", fill="x", expand=True)

        ttk.Button(btn_row, text="Cancel", command=_close_req_form, style="Req.Primary.TButton", width=12).pack(side="right")
        ttk.Button(btn_row, text="Submit", command=on_submit, style="Req.Primary.TButton", width=12).pack(side="right", padx=(0, 8))
        r += 1

        # --- FINAL: size window AFTER building the form so nothing is clipped ---
        win.update_idletasks()
        sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()

        req_w = body.inner.winfo_reqwidth()
        req_h = body.inner.winfo_reqheight()

        sb_w = body.vsb.winfo_width()
        if sb_w <= 1:
            sb_w = (body.vsb.winfo_reqwidth() or 24)

        target_w = min(max(1050, req_w + sb_w + 120), sw - 60)
        target_h = min(max(780, req_h + 120), int(sh * 0.92), sh - 90)

        win.geometry(f"{target_w}x{target_h}+40+30")
        win.minsize(target_w, 780)

        # Start at top
        win.update_idletasks()
        try:
            body.canvas.yview_moveto(0.0)
        except Exception:
            pass

        # Initial focus on first Date field
        try:
            if hasattr(self, "_req_form_first_focus") and self._req_form_first_focus is not None:
                self._req_form_first_focus.focus_set()
        except Exception:
            pass

    def _open_first_attachment(self, packed: str):
        paths = split_attachments(packed)
        if not paths:
            messagebox.showinfo("No attachment", "No attachment path saved.")
            return
        p = paths[0]
        if not os.path.exists(p):
            messagebox.showerror("Missing file", f"File not found:\n{p}")
            return
        try:
            open_file_with_default_app(p)
        except Exception as e:
            messagebox.showerror("Open failed", f"Could not open file.\n\n{e}")

    def _pick_attachments_any(self, target_var: tk.StringVar):
        paths = filedialog.askopenfilenames(
            title="Select attachment(s)",
            filetypes=[("All files", "*.*")],
        )
        if not paths:
            return
        existing = split_attachments(target_var.get())
        target_var.set(join_attachments(existing + list(paths)))

    def _pick_attachments_pdf_only(self, target_var: tk.StringVar):
        paths = filedialog.askopenfilenames(
            title="Select PDF assessment(s)",
            filetypes=[("PDF files", "*.pdf")],
        )
        if not paths:
            return
        existing = split_attachments(target_var.get())
        target_var.set(join_attachments(existing + list(paths)))

    def _date_row_form(self, parent, row, label, var: tk.StringVar) -> int:
        # Date row: Entry + Pick + (MM/DD/YYYY) grouped together
        ttk.Label(parent, text=label, style="Req.Field.TLabel").grid(row=row, column=0, sticky="w", pady=4)

        rowf = ttk.Frame(parent, style="Req.App.TFrame")
        rowf.grid(row=row, column=1, sticky="w", pady=4)

        date_entry = ttk.Entry(rowf, textvariable=var, width=18, style="Req.TEntry")
        date_entry.pack(side="left")
        # remember first required date field for initial focus
        if not hasattr(self, "_req_form_first_focus") or self._req_form_first_focus is None:
            if str(label).strip().lower().startswith("date"):
                self._req_form_first_focus = date_entry

        ttk.Button(
            rowf,
            text="Pick",
            style="Req.Pick.TButton",
            command=lambda: DatePicker(self.root, var, title=label),
        ).pack(side="left", padx=(6, 0))

        # <-- THIS is the hint, now right next to Pick
        ttk.Label(rowf, text="MM/DD/YYYY", style="Req.Sub.TLabel").pack(side="left", padx=(8, 0))

        return row + 1


    # -------------------------- Queue actions (wrappers) --------------------------
    # NOTE: The actual implementations below exist as module-level functions.
    # These wrappers ensure they are callable as instance methods.

    def _get_selected_queue_req_number(self):
        return _get_selected_queue_req_number(self)

    def _update_queue_buttons_state(self):
        return _update_queue_buttons_state(self)

    def export_queue_edrl(self):
        return export_queue_edrl(self)

    def add_selected_request_to_all(self):
        return add_selected_request_to_all(self)

    def delete_selected_request(self):
        return delete_selected_request(self)

    def edit_selected_request(self):
        return edit_selected_request(self)



# -------------------------- Queue actions --------------------------
def _get_selected_queue_req_number(self):
    """Return selected request's 'EDRL Number' (REQ-####) or None."""
    try:
        sel = self.queue_tree.selection()
        if not sel:
            return None
        values = self.queue_tree.item(sel[0], "values")
        if not values:
            return None
        return str(values[0]).strip()
    except Exception:
        return None

def _update_queue_buttons_state(self):
    has_sel = self._get_selected_queue_req_number() is not None
    try:
        self.edit_req_btn.configure(state=("normal" if has_sel else "disabled"))
    except Exception:
        pass
    try:
        self.delete_req_btn.configure(state=("normal" if has_sel else "disabled"))
    except Exception:
        pass
    try:
        self.add_software_btn.configure(state="normal")
    except Exception:
        pass

def export_queue_edrl(self):
    """Export the (filtered) queue to an Excel file in an EDRL-friendly column order."""
    df = self._queue_filtered_sorted_df()
    if df is None or df.empty:
        messagebox.showinfo("Export Queue", "There are no requests to export.")
        return

    save_path = filedialog.asksaveasfilename(
        title="Export Request Queue",
        defaultextension=".xlsx",
        filetypes=[("Excel Workbook", "*.xlsx")],
    )
    if not save_path:
        return

    out = df.copy()
    cols = [c for c in self.queue_columns_display if c in out.columns]
    out = out[cols]

    try:
        with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
            out.to_excel(writer, index=False, sheet_name="Request Queue")
        messagebox.showinfo("Export Queue", f"Exported {len(out)} request(s) to:\n{save_path}")
    except Exception as e:
        messagebox.showerror("Export Failed", f"Could not export queue.\n\n{e}")

def add_selected_request_to_all(self):
    """Add Software: take the selected Request Queue item and append it into the main EDRL sheet."""
    req = self._get_selected_queue_req_number() if hasattr(self, "_get_selected_queue_req_number") else None
    if not req:
        messagebox.showinfo("Add Software", "Select a request in the queue first.")
        return

    try:
        rec_row = self.queue_df[self.queue_df["EDRL Number"].astype(str).str.strip() == str(req).strip()]
        if rec_row.empty:
            messagebox.showerror("Add Software", "Could not find the selected request in the queue.")
            return
        rec = {k: ("" if v is None else str(v)) for k, v in rec_row.iloc[0].to_dict().items()}
    except Exception as e:
        messagebox.showerror("Add Software", f"Could not load the selected request.\n\n{e}")
        return

    # Choose the main sheet (prefer a tab named 'All', else use the first sheet)
    try:
        sheet_names = list(self.sheets.keys())
        target_sheet = self._find_sheet_case_insensitive(sheet_names, "All") or (sheet_names[0] if sheet_names else None)
    except Exception:
        target_sheet = None

    if not target_sheet or target_sheet not in self.sheets:
        messagebox.showerror("Add Software", "No workbook is loaded (or no target sheet found). Load a workbook first.")
        return

    df = self.sheets[target_sheet].copy()

    # Map Queue -> Main list normalization (per your mapping)
    # Number -> EDRL number (queue stores it as 'EDRL Number')
    mapping = {
        "EDRL number": (rec.get("EDRL Number") or rec.get("number") or rec.get("id") or ""),
        "Name": rec.get("Name", ""),
        "Version": rec.get("Version", ""),
        "Type": rec.get("Type", ""),
        "Platform": rec.get("Platform", ""),
        "Description": rec.get("Description", ""),
        "Vendor": rec.get("Vendor", ""),
        "Authorization Date": rec.get("Authorization Date", ""),
        "Authorization Expires": rec.get("Authorization Expiration", rec.get("Authorization Expires", "")),
        "State": rec.get("State", ""),
        "Authorized User": rec.get("Authorized User", ""),
    }

    # Find real columns in the sheet (case-insensitive / flexible), create if missing.
    col_candidates = {
        "EDRL number": ["edrl number", "edrl#", "edrlnumber", "number", "id"],
        "Name": ["name", "product name", "software name"],
        "Version": ["version"],
        "Type": ["type"],
        "Platform": ["platform"],
        "Description": ["description"],
        "Vendor": ["vendor"],
        "Authorization Date": ["authorization date", "approved date"],
        "Authorization Expires": ["authorization expires", "authorization expiration", "authorization expiration date", "authorization expires date"],
        "State": ["state", "approved"],
        "Authorized User": ["authorized user"],
    }

    real_cols = {}
    for logical, cands in col_candidates.items():
        real = self._find_col(df, cands)
        if not real:
            # Create new column using the logical name
            real = logical
            df[real] = ""
        real_cols[logical] = real

    new_row = {real_cols[k]: v for k, v in mapping.items()}
    # Ensure all columns exist for the row
    for c in df.columns:
        new_row.setdefault(c, "")

    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

    # Persist back to the same workbook: replace ONLY the target sheet
    wb_path = (self.workbook_path.get() if hasattr(self, "workbook_path") else "")
    wb_path = (wb_path or "").strip()
    if not wb_path or not os.path.exists(wb_path):
        messagebox.showerror("Add Software", "Workbook file not found on disk. Use Browse/Load first.")
        return

    try:
        with pd.ExcelWriter(wb_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, index=False, sheet_name=target_sheet)
    except Exception as e:
        messagebox.showerror("Add Software", f"Failed to write to workbook.\n\n{e}")
        return

    # Reload workbook + refresh UI
    try:
        self.load_workbook(wb_path)
        # switch to All sheet so you can immediately see it
        try:
            self.current_sheet.set(target_sheet)
            self.on_sheet_change()
        except Exception:
            pass
        messagebox.showinfo("Add Software", f"Added '{mapping.get('Name','')}' to '{target_sheet}'.")
        # Optional: remove from queue after successful add (with confirmation)
        try:
            prompt = (
                f"'{req}' was added to '{target_sheet}'.\n\n"
                "Remove it from the Request Queue?"
            )
            if messagebox.askyesno("Remove from Queue", prompt):
                if "EDRL Number" in self.queue_df.columns:
                    self.queue_df = self.queue_df[self.queue_df["EDRL Number"].astype(str).str.strip() != str(req).strip()].copy()
                self.refresh_queue_table()
                self._update_queue_buttons_state()
        except Exception:
            pass
    except Exception as e:
        messagebox.showwarning("Add Software", f"Added, but could not refresh UI automatically.\n\n{e}")

def delete_selected_request(self):
    req = self._get_selected_queue_req_number()
    if not req:
        messagebox.showinfo("Delete Request", "Select a request in the queue first.")
        return

    if messagebox.askyesno("Delete Request", f"Delete request '{req}' from the queue?"):
        try:
            if "EDRL Number" in self.queue_df.columns:
                self.queue_df = self.queue_df[self.queue_df["EDRL Number"].astype(str).str.strip() != req].copy()
            self.refresh_queue_table()
            self._update_queue_buttons_state()
        except Exception as e:
            messagebox.showerror("Delete Failed", f"Could not delete request.\n\n{e}")

def edit_selected_request(self):
    req = self._get_selected_queue_req_number()
    if not req:
        messagebox.showinfo("Edit Request", "Select a request in the queue first.")
        return

    try:
        rec_row = self.queue_df[self.queue_df["EDRL Number"].astype(str).str.strip() == req]
        if rec_row.empty:
            messagebox.showerror("Edit Request", "Could not find the selected request in the queue.")
            return
        rec = rec_row.iloc[0].to_dict()
    except Exception as e:
        messagebox.showerror("Edit Request", f"Could not load request for editing.\n\n{e}")
        return

    self.open_request_form(prefill=rec, edit_req_number=req)





def _update_results_buttons_state(self):
    """UI-only: enable/disable Results buttons based on selection/content."""
    try:
        has_rows = bool(self.results_tree.get_children())
        has_sel = bool(self.results_tree.selection())

        if hasattr(self, "results_clear_btn") and self.results_clear_btn:
            self.results_clear_btn.configure(state=("normal" if has_rows else "disabled"))

        if hasattr(self, "results_delete_btn") and self.results_delete_btn:
            self.results_delete_btn.configure(state=("normal" if has_sel else "disabled"))
    except Exception:
        pass

def delete_selected_result(self):
    """UI-only: remove selected row(s) from the Results table (does not modify Excel)."""
    try:
        sel = self.results_tree.selection()
        if not sel:
            self._update_results_buttons_state()
            return

        for iid in sel:
            self.results_tree.delete(iid)

        remaining = len(self.results_tree.get_children())
        try:
            self.status_var.set(f"Deleted selected row(s) from Results (UI only). Rows remaining: {remaining}")
        except Exception:
            pass

        self._update_results_buttons_state()
    except Exception as e:
        try:
            messagebox.showerror("Delete Failed", f"Could not delete selected row(s).\n\n{e}")
        except Exception:
            pass

def clear_results_table(self):
    """UI-only: clear all rows from the Results table (does not modify Excel)."""
    try:
        children = self.results_tree.get_children()
        if children:
            self.results_tree.delete(*children)

        try:
            self.status_var.set("Results cleared (UI only).")
        except Exception:
            pass

        self._update_results_buttons_state()
    except Exception as e:
        try:
            messagebox.showerror("Clear Failed", f"Could not clear results.\n\n{e}")
        except Exception:
            pass

# --- Bind queue action functions as EDRLSearchGUI methods (in case file indentation changes) ---
EDRLSearchGUI.add_selected_request_to_all = add_selected_request_to_all
EDRLSearchGUI.delete_selected_request = delete_selected_request
EDRLSearchGUI.edit_selected_request = edit_selected_request
EDRLSearchGUI.export_queue_edrl = export_queue_edrl
EDRLSearchGUI._update_queue_buttons_state = _update_queue_buttons_state
EDRLSearchGUI._update_results_buttons_state = _update_results_buttons_state
EDRLSearchGUI.delete_selected_result = delete_selected_result
EDRLSearchGUI.clear_results_table = clear_results_table
def main():
    root = tk.Tk()
    EDRLSearchGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()